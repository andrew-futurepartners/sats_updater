import io
import os
import re
import math
import pandas as pd
import streamlit as st

# =========================
# Config
# =========================
# Set this to your repo-relative path for DataMapping.xlsx
MAPPING_PATH = os.environ.get("MAPPING_PATH", "DataMapping.xlsx")


# =========================
# Column resolution helpers
# =========================
def build_column_index(df: pd.DataFrame):
    exact = set(df.columns)
    lower_map = {}
    collisions = set()
    for c in df.columns:
        key = c.lower()
        if key in lower_map and lower_map[key] != c:
            collisions.add(key)
        else:
            lower_map[key] = c
    for k in collisions:
        lower_map.pop(k, None)
    return {"exact": exact, "lower_map": lower_map, "collisions": collisions}

def resolve_col(name: str, col_index):
    if name in col_index["exact"]:
        return name
    return col_index["lower_map"].get(name.lower(), None)


# =====================
# Mapping file parsing
# =====================
def parse_data_mapping(mapping_path):
    df = pd.read_excel(mapping_path)
    df = df.dropna(subset=["FINAL"])

    standard_mappings = {}
    pattern_mappings_1 = []
    pattern_mappings_2 = []
    final_column_order = df["FINAL"].dropna().tolist()

    for _, row in df.iterrows():
        final = str(row["FINAL"]).strip()
        if final not in final_column_order:
            final_column_order.append(final)

        o1 = row.get("ORIGINAL_1")
        if pd.notna(o1) and not str(o1).startswith("["):
            o1 = str(o1).strip()
            if re.search(r"(LrNr|\{N\})", o1, flags=re.IGNORECASE):
                pattern_mappings_1.append({"pattern": o1, "final_column": final})
            else:
                standard_mappings[o1] = final

        o2 = row.get("ORIGINAL_2")
        if pd.notna(o2) and not str(o2).startswith("["):
            o2 = str(o2).strip()
            if re.search(r"(LrNr|\{N\})", o2, flags=re.IGNORECASE):
                pattern_mappings_2.append({"pattern": o2, "final_column": final})
            else:
                standard_mappings[o2] = final

    return standard_mappings, pattern_mappings_1, pattern_mappings_2, final_column_order


# =======================
# Destination utilities
# =======================
def extract_destination_codes(df: pd.DataFrame):
    dest_pattern = re.compile(r"_lr(\d+)", re.IGNORECASE)
    dest_nums = set()
    for col in df.columns:
        m = dest_pattern.search(col)
        if m:
            dest_nums.add(int(m.group(1)))
    return [f"lr{n}" for n in sorted(dest_nums)]

def expand_pattern_columns(pattern_mappings, dest_codes):
    expanded = {}
    for entry in pattern_mappings:
        pattern = entry["pattern"].strip()
        final_col = entry["final_column"]
        m = re.match(r"([A-Za-z]\d[\w]*)_Lr(?:Nr|\{N\})(?:r(\d+))?$", pattern, flags=re.IGNORECASE)
        if not m:
            continue
        base_q = m.group(1)
        r_tail = m.group(2)
        for dest in dest_codes:
            full_col = f"{base_q}_{dest}"
            if r_tail:
                full_col += f"r{r_tail}"
            expanded[full_col.lower()] = final_col
    return expanded


# ===========================
# SATS file loading utility
# ===========================
def load_sats_dataframe(uploaded_file, sheet_name=None):
    xls = pd.ExcelFile(uploaded_file)
    sheets = xls.sheet_names if sheet_name in (None, "",) else [sheet_name]
    for sn in sheets:
        df_try = pd.read_excel(uploaded_file, sheet_name=sn)
        if any(str(c).strip().lower() == "markers" for c in df_try.columns):
            return df_try, sn
    first_sn = sheets[0]
    return pd.read_excel(uploaded_file, sheet_name=first_sn), first_sn


# ==============================
# Build mapping and data bundle
# ==============================
def build_full_mapping(mapping_path, sats_file, sats_filename, sheet_name=None, wave_label="", year_label=""):
    standard_mappings, patt1, patt2, final_column_order = parse_data_mapping(mapping_path)
    df, chosen_sheet = load_sats_dataframe(sats_file, sheet_name=sheet_name)
    col_index = build_column_index(df)
    dest_codes = extract_destination_codes(df)
    expanded_1 = expand_pattern_columns(patt1, dest_codes)
    expanded_2 = expand_pattern_columns(patt2, dest_codes)

    full_mapping = {}
    full_mapping.update({k: v for k, v in standard_mappings.items()})
    full_mapping.update(expanded_1)
    full_mapping.update(expanded_2)

    # Wave and year come from user input. Year optional, infer from wave if needed.
    survey_year = year_label
    if not survey_year:
        m = re.search(r"\b(20\d{2})\b", wave_label)
        survey_year = m.group(1) if m else ""

    for extra in ["Wave", "HIDE Help", "CITY_EVAL"]:
        if extra not in final_column_order:
            final_column_order.append(extra)

    return full_mapping, dest_codes, df, list(df.columns), final_column_order, (expanded_1, expanded_2), wave_label, survey_year, col_index


# =========================
# Markers parsing helpers
# =========================
def marker_label_from_blob(markers_str: str, label_type: str, n: int):
    if not isinstance(markers_str, str) or not markers_str:
        return pd.NA
    pat = rf"(?:^|,)\s*/?[^,]*\b{label_type}\s*0*{n}/([^,]+)"
    m = re.search(pat, markers_str, flags=re.IGNORECASE)
    return m.group(1).strip() if m else pd.NA


# ==================
# Reshape function
# ==================
def reshape_sats_data(df, full_map, dest_codes, final_column_order, original_to_final_groups, wave_label="", survey_year="", col_index=None):
    if col_index is None:
        col_index = build_column_index(df)

    mapping1, mapping2 = original_to_final_groups  # mapping1 = CITY group, mapping2 = STATE group
    lower_to_actual = {c.lower(): c for c in df.columns}
    markers_blob_col = resolve_col("markers", col_index)

    static_keys = [k for k in full_map.keys() if not re.search(r"lr\d+", str(k), flags=re.IGNORECASE)]
    city_records = []
    state_records = []

    for _, row in df.iterrows():
        found_dests_1 = set()
        found_dests_2 = set()

        for original_lower in mapping1.keys():
            actual = lower_to_actual.get(original_lower)
            if actual is not None and pd.notna(row.get(actual, None)):
                m = re.search(r"(lr\d+)", original_lower, flags=re.IGNORECASE)
                if m:
                    found_dests_1.add(m.group(1).lower())

        for original_lower in mapping2.keys():
            actual = lower_to_actual.get(original_lower)
            if actual is not None and pd.notna(row.get(actual, None)):
                m = re.search(r"(lr\d+)", original_lower, flags=re.IGNORECASE)
                if m:
                    found_dests_2.add(m.group(1).lower())

        static_row = {}
        for src in static_keys:
            actual_src = resolve_col(src, col_index)
            if actual_src is not None:
                val = row.get(actual_src, None)
                if pd.notna(val):
                    static_row[full_map[src]] = val

        static_row["Wave"] = wave_label or pd.NA
        static_row["HIDE Help"] = survey_year or pd.NA

        markers_str = str(row.get(markers_blob_col, "")) if markers_blob_col else ""

        # CITY first
        for dest in sorted(found_dests_1, key=lambda s: int(re.search(r"\d+", s).group(0))):
            row_out = static_row.copy()
            for original_lower, final in mapping1.items():
                target_lower = re.sub(r"lr\d+", dest, original_lower, flags=re.IGNORECASE)
                actual = lower_to_actual.get(target_lower)
                if actual is not None:
                    v = row.get(actual, None)
                    if pd.notna(v):
                        row_out[final] = v
            row_out["CITY_EVAL"] = marker_label_from_blob(markers_str, "Destination", 1)
            for c in final_column_order:
                row_out.setdefault(c, pd.NA)
            city_records.append(row_out)

        # STATE second
        for dest in sorted(found_dests_2, key=lambda s: int(re.search(r"\d+", s).group(0))):
            row_out = static_row.copy()
            for original_lower, final in mapping2.items():
                target_lower = re.sub(r"lr\d+", dest, original_lower, flags=re.IGNORECASE)
                actual = lower_to_actual.get(target_lower)
                if actual is not None:
                    v = row.get(actual, None)
                    if pd.notna(v):
                        row_out[final] = v
            row_out["CITY_EVAL"] = marker_label_from_blob(markers_str, "State", 1)
            for c in final_column_order:
                row_out.setdefault(c, pd.NA)
            state_records.append(row_out)

    out_df = pd.DataFrame(city_records + state_records)
    out_df = out_df.reindex(columns=list(dict.fromkeys(final_column_order)))
    return out_df


# ============================
# Weight helpers for big files
# ============================
def _normalize_wave(s: str) -> str:
    if s is None:
        return ""
    x = str(s).strip().lower().replace("_", " ")
    x = re.sub(r"\s+", " ", x)
    return x

def read_weight_for_wave(uploaded_file, wave_label, sheet_name=None, weight_header_guess=None, wave_header_guess="Wave"):
    """
    Stream only the Wave and Weight columns from a very large Excel or CSV.
    Filter rows where Wave equals the typed wave_label.
    Return a Series of weights in file order for that wave.
    """
    norm_target = _normalize_wave(wave_label or "")

    # CSV path
    name = getattr(uploaded_file, "name", "").lower()
    if name.endswith(".csv"):
        uploaded_file.seek(0)
        head = pd.read_csv(uploaded_file, nrows=0)
        cols = head.columns.tolist()
        # pick wave and weight columns
        wave_col = None
        weight_col = None

        # exact guesses first
        for c in cols:
            if wave_header_guess and str(c).strip() == wave_header_guess:
                wave_col = c
            if weight_header_guess and str(c).strip() == weight_header_guess:
                weight_col = c
        # fallbacks
        if wave_col is None:
            for c in cols:
                if str(c).strip().lower() == "wave":
                    wave_col = c
                    break
        if weight_col is None:
            for cand in ["Weight", "WEIGHT", "weight", "wt", "WT", "WGT", "wgt"]:
                for c in cols:
                    if str(c).strip() == cand:
                        weight_col = c
                        break
                if weight_col is not None:
                    break

        if wave_col is None or weight_col is None:
            raise ValueError("Could not find Wave or Weight columns in CSV")

        uploaded_file.seek(0)
        df = pd.read_csv(uploaded_file, usecols=[wave_col, weight_col])
        mask = df[wave_col].astype(str).map(_normalize_wave) == norm_target
        return df.loc[mask, weight_col].reset_index(drop=True)

    # Excel path, stream with openpyxl
    try:
        from openpyxl import load_workbook
    except Exception:
        # fallback, read only two columns using pandas
        df = pd.read_excel(uploaded_file, sheet_name=sheet_name)
        # locate columns
        def pick(cols):
            wave_col = None
            weight_col = None
            for c in cols:
                if wave_header_guess and str(c).strip() == wave_header_guess:
                    wave_col = c
                if weight_header_guess and str(c).strip() == weight_header_guess:
                    weight_col = c
            if wave_col is None:
                for c in cols:
                    if str(c).strip().lower() == "wave":
                        wave_col = c
                        break
            if weight_col is None:
                for cand in ["Weight", "WEIGHT", "weight", "wt", "WT", "WGT", "wgt"]:
                    for c in cols:
                        if str(c).strip() == cand:
                            weight_col = c
                            break
                    if weight_col is not None:
                        break
            return wave_col, weight_col

        wave_col, weight_col = pick(df.columns)
        if wave_col is None or weight_col is None:
            raise ValueError("Could not find Wave or Weight columns in Excel")
        mask = df[wave_col].astype(str).map(_normalize_wave) == norm_target
        return df.loc[mask, weight_col].reset_index(drop=True)

    uploaded_file.seek(0)
    wb = load_workbook(uploaded_file, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.worksheets[0]

    # header
    header_cells = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [str(c.value).strip() if c.value is not None else "" for c in header_cells]

    # find indices
    wave_idx = None
    weight_idx = None

    # exact guesses
    if wave_header_guess:
        for j, h in enumerate(headers):
            if h == wave_header_guess:
                wave_idx = j
                break
    if weight_header_guess:
        for j, h in enumerate(headers):
            if h == weight_header_guess:
                weight_idx = j
                break

    # fallbacks
    if wave_idx is None:
        for j, h in enumerate(headers):
            if h.strip().lower() == "wave":
                wave_idx = j
                break
    if weight_idx is None:
        for cand in ["Weight", "WEIGHT", "weight", "wt", "WT", "WGT", "wgt"]:
            for j, h in enumerate(headers):
                if h == cand:
                    weight_idx = j
                    break
            if weight_idx is not None:
                break

    if wave_idx is None or weight_idx is None:
        wb.close()
        raise ValueError("Could not find Wave or Weight columns in Excel")

    weights = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        wv = row[wave_idx] if wave_idx < len(row) else None
        if _normalize_wave(wv) == norm_target:
            wt = row[weight_idx] if weight_idx < len(row) else None
            weights.append(wt)

    wb.close()
    return pd.Series(weights)


# ===============
# Streamlit UI
# ===============
st.title("SATS+ Reshaper")
st.caption("Upload SATS data. The app uses DataMapping.xlsx from the repo. Optional: upload a large SATS file with Weight and we will pull only the current Wave and stack it.")

with st.form("sats_form"):
    sats_file = st.file_uploader("SATS datafile (xlsx or xls). Must include a 'markers' column", type=["xlsx", "xls"])
    sats_sheet = st.text_input("SATS sheet name (optional)", value="")
    wave_input = st.text_input("Current Wave (required, for example 'June 2025')", value="")
    year_input = st.text_input("Survey year (optional, 4 digits)", value="")
    st.markdown("---")
    st.subheader("Optional weight injection from a big SATS file")
    weight_file = st.file_uploader("Large SATS file containing Wave and Weight (CSV or Excel). We will stream only those two columns", type=["xlsx", "xls", "csv"])
    weight_sheet = st.text_input("Weight sheet name (optional)", value="")
    weight_header = st.text_input("Weight column header (optional, defaults to Weight)", value="")
    wave_header = st.text_input("Wave column header in the weight file (optional, defaults to Wave)", value="")
    submitted = st.form_submit_button("Process")

if submitted:
    if not sats_file:
        st.error("Please upload the SATS datafile.")
        st.stop()
    if not wave_input.strip():
        st.error("Please type the current Wave, for example 'June 2025'.")
        st.stop()
    if not os.path.exists(MAPPING_PATH):
        st.error(f"Mapping file not found at '{MAPPING_PATH}'. Set MAPPING_PATH env var or place DataMapping.xlsx in the app folder.")
        st.stop()

    # Build mapping and reshape
    fm, dest_codes, df_raw, cols_raw, final_order, groups, wave_label, survey_year, col_index = build_full_mapping(
        MAPPING_PATH, sats_file, getattr(sats_file, "name", ""), sheet_name=sats_sheet or None,
        wave_label=wave_input.strip(), year_label=year_input.strip()
    )

    out_df = reshape_sats_data(df_raw, fm, dest_codes, final_order, groups, wave_label, survey_year, col_index)

    # Optional weight stacking filtered by Wave
    if weight_file is not None:
        try:
            weights = read_weight_for_wave(
                weight_file, wave_label,
                sheet_name=weight_sheet or None,
                weight_header_guess=weight_header or None,
                wave_header_guess=wave_header or "Wave",
            ).reset_index(drop=True)

            n_resp = len(df_raw)
            if len(weights) == 0:
                st.warning("No weights found for the Wave you typed. Check the Wave text and headers.")
            elif len(weights) < n_resp:
                st.warning(f"Weight rows {len(weights)} are fewer than datafile rows {n_resp}. Will truncate to the shorter length.")
                n = len(weights)
            else:
                n = n_resp

            # Stack weights into CITY then STATE
            stacked = pd.concat([weights.iloc[:n].reset_index(drop=True),
                                 weights.iloc[:n].reset_index(drop=True)], ignore_index=True)

            # Align length
            if len(stacked) != len(out_df):
                if len(stacked) < len(out_df):
                    stacked = pd.concat([stacked, pd.Series([None] * (len(out_df) - len(stacked)))], ignore_index=True)
                else:
                    stacked = stacked.iloc[:len(out_df)].reset_index(drop=True)

            out_df.insert(len(out_df.columns), "WEIGHT", stacked)
            st.success("WEIGHT column added by file order for the selected Wave.")
        except Exception as e:
            st.error(f"Could not extract WEIGHT: {e}")

    # Prepare download
    out_name = f"{wave_label} SATS+ Output.xlsx" if wave_label else "SATS_final_output.xlsx"
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter") as xw:
        out_df.to_excel(xw, index=False, sheet_name="SATS+")
    st.download_button("Download Output Excel", data=buf.getvalue(), file_name=out_name, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # Quick previews
    st.write("Preview of output")
    st.dataframe(out_df.head(10))
    st.write("Tail of output")
    st.dataframe(out_df.tail(10))
